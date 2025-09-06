#!/usr/bin/env python3
"""
Calendar Integration System with Calendly-style functionality
Manages appointment scheduling, calendar slots, and Excel export integration
"""

import sqlite3
import pandas as pd
from datetime import datetime, timedelta, time
import json
import uuid
import os
from typing import Dict, List, Optional, Tuple
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import calendar

class CalendarIntegration:
    """Enhanced calendar integration with Calendly-style features"""
    
    def __init__(self):
        self.db_path = "data/medical_scheduler.db"
        self.excel_export_path = "exports"
        self.calendar_data_path = "data/calendar_data.json"
        
        # Ensure export directory exists
        os.makedirs(self.excel_export_path, exist_ok=True)
        
        # Initialize calendar configuration
        self.working_hours = {
            'start_time': time(9, 0),  # 9:00 AM
            'end_time': time(17, 0),   # 5:00 PM
            'lunch_start': time(12, 0), # 12:00 PM
            'lunch_end': time(13, 0),   # 1:00 PM
        }
        
        self.slot_duration = 30  # minutes
        self.buffer_time = 15    # minutes between appointments
        
        # Days of week (0=Monday, 6=Sunday)
        self.working_days = [0, 1, 2, 3, 4]  # Monday to Friday
        
    def get_db_connection(self):
        """Get database connection"""
        return sqlite3.connect(self.db_path)
    
    def generate_available_slots(self, doctor_id: str, start_date: str, days_ahead: int = 14) -> List[Dict]:
        """
        Generate available appointment slots for a doctor (Calendly-style)
        
        Args:
            doctor_id: Doctor's ID
            start_date: Start date (YYYY-MM-DD format)
            days_ahead: Number of days to generate slots for
            
        Returns:
            List of available slots
        """
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            slots = []
            
            # Get existing appointments for this doctor
            conn = self.get_db_connection()
            existing_query = """
                SELECT appointment_date, appointment_time 
                FROM appointments 
                WHERE doctor_id = ? AND appointment_date >= ?
            """
            existing_appointments = pd.read_sql_query(
                existing_query, conn, params=[doctor_id, start_date]
            )
            conn.close()
            
            # Convert existing appointments to set for quick lookup
            booked_slots = set()
            for _, appt in existing_appointments.iterrows():
                booked_slots.add(f"{appt['appointment_date']}_{appt['appointment_time']}")
            
            # Generate slots for each day
            for day_offset in range(days_ahead):
                current_date = start_dt + timedelta(days=day_offset)
                
                # Skip weekends if not in working days
                if current_date.weekday() not in self.working_days:
                    continue
                
                # Skip past dates
                if current_date.date() < datetime.now().date():
                    continue
                
                date_str = current_date.strftime('%Y-%m-%d')
                day_slots = self._generate_day_slots(date_str, booked_slots)
                slots.extend(day_slots)
            
            return slots
            
        except Exception as e:
            print(f"Error generating slots: {e}")
            return []
    
    def _generate_day_slots(self, date_str: str, booked_slots: set) -> List[Dict]:
        """Generate available slots for a specific day"""
        slots = []
        current_time = datetime.combine(
            datetime.strptime(date_str, '%Y-%m-%d').date(),
            self.working_hours['start_time']
        )
        end_time = datetime.combine(
            datetime.strptime(date_str, '%Y-%m-%d').date(),
            self.working_hours['end_time']
        )
        lunch_start = datetime.combine(
            datetime.strptime(date_str, '%Y-%m-%d').date(),
            self.working_hours['lunch_start']
        )
        lunch_end = datetime.combine(
            datetime.strptime(date_str, '%Y-%m-%d').date(),
            self.working_hours['lunch_end']
        )
        
        while current_time + timedelta(minutes=self.slot_duration) <= end_time:
            # Skip lunch hour
            if lunch_start <= current_time < lunch_end:
                current_time = lunch_end
                continue
            
            time_str = current_time.strftime('%H:%M')
            slot_key = f"{date_str}_{time_str}"
            
            # Check if slot is not booked
            if slot_key not in booked_slots:
                slots.append({
                    'date': date_str,
                    'time': time_str,
                    'datetime': current_time.isoformat(),
                    'day_name': current_time.strftime('%A'),
                    'formatted_date': current_time.strftime('%B %d, %Y'),
                    'formatted_time': current_time.strftime('%I:%M %p'),
                    'slot_id': str(uuid.uuid4()),
                    'available': True
                })
            
            current_time += timedelta(minutes=self.slot_duration + self.buffer_time)
        
        return slots
    
    def get_doctor_availability(self, doctor_id: str, date_range: int = 14) -> Dict:
        """Get comprehensive availability for a doctor"""
        try:
            start_date = datetime.now().strftime('%Y-%m-%d')
            available_slots = self.generate_available_slots(doctor_id, start_date, date_range)
            
            # Group slots by date
            slots_by_date = {}
            for slot in available_slots:
                date = slot['date']
                if date not in slots_by_date:
                    slots_by_date[date] = []
                slots_by_date[date].append(slot)
            
            # Get doctor info with proper parameter order
            conn = self.get_db_connection()
            doctor_query = "SELECT * FROM doctors WHERE doctor_id = ?"
            doctor_info = pd.read_sql_query(doctor_query, conn, params=[doctor_id])
            conn.close()
            
            if doctor_info.empty:
                return {'error': f'Doctor with ID {doctor_id} not found'}
            
            doctor = doctor_info.iloc[0].to_dict()
            
            return {
                'doctor': doctor,
                'total_slots': len(available_slots),
                'slots_by_date': slots_by_date,
                'date_range': f"{start_date} to {(datetime.now() + timedelta(days=date_range)).strftime('%Y-%m-%d')}",
                'working_hours': f"{self.working_hours['start_time'].strftime('%I:%M %p')} - {self.working_hours['end_time'].strftime('%I:%M %p')}",
                'slot_duration': f"{self.slot_duration} minutes"
            }
            
        except Exception as e:
            print(f"Detailed error in get_doctor_availability: {e}")
            return {'error': f'Error getting availability: {e}'}
    
    def book_calendly_slot(self, slot_data: Dict) -> Dict:
        """
        Book an appointment slot (Calendly-style booking)
        
        Args:
            slot_data: Dictionary containing booking information
            
        Returns:
            Booking confirmation data
        """
        try:
            # Generate appointment ID
            appointment_id = str(uuid.uuid4())[:8].upper()
            
            # Save to database
            conn = self.get_db_connection()
            cursor = conn.cursor()
            
            # Insert appointment
            cursor.execute("""
                INSERT INTO appointments (
                    appointment_id, patient_id, doctor_id, appointment_date, 
                    appointment_time, duration_minutes, appointment_type, 
                    status, created_at, booking_method
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                appointment_id,
                slot_data['patient_id'],
                slot_data['doctor_id'],
                slot_data['date'],
                slot_data['time'],
                slot_data.get('duration', self.slot_duration),
                slot_data.get('appointment_type', 'consultation'),
                'confirmed',
                datetime.now().isoformat(),
                'calendar_integration'
            ))
            
            conn.commit()
            conn.close()
            
            # Generate calendar data for Excel export
            calendar_data = self._create_calendar_entry(appointment_id, slot_data)
            
            # Export to Excel
            excel_file = self.export_appointment_to_calendar_excel(appointment_id, calendar_data)
            
            return {
                'success': True,
                'appointment_id': appointment_id,
                'booking_confirmation': calendar_data,
                'excel_export': excel_file,
                'calendar_link': self._generate_calendar_link(calendar_data),
                'message': f'Appointment {appointment_id} successfully booked!'
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Booking failed: {e}'
            }
    
    def _create_calendar_entry(self, appointment_id: str, slot_data: Dict) -> Dict:
        """Create calendar entry data"""
        try:
            # Get additional details
            conn = self.get_db_connection()
            
            # Get patient info
            patient_query = "SELECT * FROM patients WHERE patient_id = ?"
            patient_info = pd.read_sql_query(patient_query, conn, params=[slot_data['patient_id']])
            
            # Get doctor info
            doctor_query = "SELECT * FROM doctors WHERE doctor_id = ?"
            doctor_info = pd.read_sql_query(doctor_query, conn, params=[slot_data['doctor_id']])
            
            conn.close()
            
            # Safely get data with defaults
            patient = {}
            doctor = {}
            
            if not patient_info.empty:
                patient = patient_info.iloc[0].to_dict()
            
            if not doctor_info.empty:
                doctor = doctor_info.iloc[0].to_dict()
            
            # Create datetime objects
            appointment_datetime = datetime.strptime(
                f"{slot_data['date']} {slot_data['time']}", 
                "%Y-%m-%d %H:%M"
            )
            end_datetime = appointment_datetime + timedelta(minutes=self.slot_duration)
            
            calendar_entry = {
                'appointment_id': appointment_id,
                'title': f"Medical Appointment - {patient.get('first_name', 'Unknown')} {patient.get('last_name', 'Patient')}",
                'doctor': f"Dr. {doctor.get('doctor_name', 'Unknown Doctor')}",
                'specialty': doctor.get('specialty', 'General Medicine'),
                'patient_name': f"{patient.get('first_name', 'Unknown')} {patient.get('last_name', 'Patient')}",
                'patient_email': patient.get('email', 'no-email@example.com'),
                'patient_phone': patient.get('phone', '(000) 000-0000'),
                'start_datetime': appointment_datetime.isoformat(),
                'end_datetime': end_datetime.isoformat(),
                'date': slot_data['date'],
                'time': slot_data['time'],
                'duration': f"{self.slot_duration} minutes",
                'location': "MediCare Allergy & Wellness Center",
                'description': f"Appointment with Dr. {doctor.get('doctor_name', 'Unknown Doctor')} ({doctor.get('specialty', 'General Medicine')})",
                'status': 'confirmed',
                'created_at': datetime.now().isoformat(),
                'timezone': 'Local Time'
            }
            
            return calendar_entry
            
        except Exception as e:
            print(f"Error creating calendar entry: {e}")
            # Return a basic calendar entry as fallback
            return {
                'appointment_id': appointment_id,
                'title': "Medical Appointment",
                'doctor': "Dr. Unknown",
                'specialty': "General Medicine",
                'patient_name': "Unknown Patient",
                'patient_email': 'no-email@example.com',
                'patient_phone': '(000) 000-0000',
                'start_datetime': datetime.now().isoformat(),
                'end_datetime': (datetime.now() + timedelta(minutes=30)).isoformat(),
                'date': slot_data.get('date', '2025-09-06'),
                'time': slot_data.get('time', '09:00'),
                'duration': "30 minutes",
                'location': "MediCare Allergy & Wellness Center",
                'description': "Medical Appointment",
                'status': 'confirmed',
                'created_at': datetime.now().isoformat(),
                'timezone': 'Local Time'
            }
    
    def export_appointment_to_calendar_excel(self, appointment_id: str, calendar_data: Dict) -> str:
        """Export appointment to Excel with calendar-style formatting"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"calendar_appointment_{appointment_id}_{timestamp}.xlsx"
            filepath = os.path.join(self.excel_export_path, filename)
            
            # Create workbook and worksheet
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = "Calendar Appointment"
            
            # Define styles
            header_font = Font(bold=True, size=14, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            subheader_font = Font(bold=True, size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Title
            ws.merge_cells('A1:F1')
            ws['A1'] = "📅 MEDICAL APPOINTMENT CALENDAR"
            ws['A1'].font = header_font
            ws['A1'].fill = header_fill
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Appointment details
            row = 3
            
            # Basic Info Section
            ws[f'A{row}'] = "APPOINTMENT DETAILS"
            ws[f'A{row}'].font = subheader_font
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
            
            details = [
                ("Appointment ID:", calendar_data['appointment_id']),
                ("Patient Name:", calendar_data['patient_name']),
                ("Doctor:", calendar_data['doctor']),
                ("Specialty:", calendar_data['specialty']),
                ("Date:", datetime.fromisoformat(calendar_data['start_datetime']).strftime('%A, %B %d, %Y')),
                ("Time:", f"{calendar_data['time']} ({calendar_data['duration']})"),
                ("Location:", calendar_data['location']),
                ("Status:", calendar_data['status'].upper())
            ]
            
            for label, value in details:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            row += 1
            
            # Contact Information
            ws[f'A{row}'] = "CONTACT INFORMATION"
            ws[f'A{row}'].font = subheader_font
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
            
            contact_info = [
                ("Patient Email:", calendar_data['patient_email']),
                ("Patient Phone:", calendar_data['patient_phone']),
                ("Clinic Phone:", "(555) 123-4567"),
                ("Clinic Email:", "appointments@medicare.com")
            ]
            
            for label, value in contact_info:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
            
            row += 2
            
            # Calendar Instructions
            ws[f'A{row}'] = "CALENDAR INTEGRATION"
            ws[f'A{row}'].font = subheader_font
            ws.merge_cells(f'A{row}:F{row}')
            row += 1
            
            instructions = [
                "• Add this appointment to your calendar",
                "• Set reminder 24 hours before appointment",
                "• Arrive 15 minutes early for check-in",
                "• Bring valid ID and insurance card",
                "• Complete intake forms if new patient"
            ]
            
            for instruction in instructions:
                ws[f'A{row}'] = instruction
                row += 1
            
            row += 2
            
            # Quick Calendar View (Mini Calendar)
            self._add_mini_calendar(ws, row, calendar_data)
            
            # Apply borders and formatting
            for row_cells in ws.iter_rows(min_row=3, max_row=row+10, min_col=1, max_col=6):
                for cell in row_cells:
                    cell.border = border
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            
            # Save workbook
            workbook.save(filepath)
            
            return filename
            
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            return None
    
    def _add_mini_calendar(self, ws, start_row: int, calendar_data: Dict):
        """Add a mini calendar view to the Excel sheet"""
        try:
            appointment_date = datetime.fromisoformat(calendar_data['start_datetime'])
            year = appointment_date.year
            month = appointment_date.month
            
            # Calendar header
            ws[f'A{start_row}'] = "CALENDAR VIEW"
            ws[f'A{start_row}'].font = Font(bold=True, size=12)
            ws.merge_cells(f'A{start_row}:G{start_row}')
            start_row += 1
            
            # Month and year
            month_name = calendar.month_name[month]
            ws[f'A{start_row}'] = f"{month_name} {year}"
            ws[f'A{start_row}'].font = Font(bold=True)
            ws.merge_cells(f'A{start_row}:G{start_row}')
            start_row += 1
            
            # Day headers
            days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
            for i, day in enumerate(days):
                ws.cell(row=start_row, column=i+1, value=day)
                ws.cell(row=start_row, column=i+1).font = Font(bold=True)
            start_row += 1
            
            # Calendar grid
            cal = calendar.monthcalendar(year, month)
            for week in cal:
                for i, day in enumerate(week):
                    if day == 0:
                        ws.cell(row=start_row, column=i+1, value="")
                    else:
                        cell = ws.cell(row=start_row, column=i+1, value=day)
                        # Highlight appointment day
                        if day == appointment_date.day:
                            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                            cell.font = Font(bold=True, color="FF0000")
                start_row += 1
                
        except Exception as e:
            print(f"Error adding mini calendar: {e}")
    
    def _generate_calendar_link(self, calendar_data: Dict) -> str:
        """Generate calendar link for external calendar applications"""
        # Simple calendar link (can be enhanced for specific calendar apps)
        start_time = calendar_data['start_datetime'].replace(':', '%3A')
        end_time = calendar_data['end_datetime'].replace(':', '%3A')
        
        # Google Calendar link format
        google_calendar_link = (
            f"https://calendar.google.com/calendar/render?action=TEMPLATE"
            f"&text={calendar_data['title'].replace(' ', '%20')}"
            f"&dates={start_time.replace('-', '').replace(':', '')}Z/"
            f"{end_time.replace('-', '').replace(':', '')}Z"
            f"&details={calendar_data['description'].replace(' ', '%20')}"
            f"&location={calendar_data['location'].replace(' ', '%20')}"
        )
        
        return google_calendar_link
    
    def get_all_appointments_calendar(self, start_date: str = None, end_date: str = None) -> Dict:
        """Get all appointments in calendar format"""
        try:
            conn = self.get_db_connection()
            
            # Build query
            base_query = """
                SELECT a.*, p.first_name, p.last_name, p.email, p.phone,
                       d.doctor_name, d.specialty
                FROM appointments a
                LEFT JOIN patients p ON a.patient_id = p.patient_id
                LEFT JOIN doctors d ON a.doctor_id = d.doctor_id
            """
            
            params = []
            if start_date and end_date:
                base_query += " WHERE a.appointment_date BETWEEN ? AND ?"
                params = [start_date, end_date]
            elif start_date:
                base_query += " WHERE a.appointment_date >= ?"
                params = [start_date]
            
            base_query += " ORDER BY a.appointment_date, a.appointment_time"
            
            appointments_df = pd.read_sql_query(base_query, conn, params=params)
            conn.close()
            
            # Convert to calendar format
            calendar_events = []
            for _, appt in appointments_df.iterrows():
                # Create datetime
                appt_datetime = datetime.strptime(
                    f"{appt['appointment_date']} {appt['appointment_time']}", 
                    "%Y-%m-%d %H:%M"
                )
                end_datetime = appt_datetime + timedelta(minutes=appt.get('duration_minutes', 30))
                
                event = {
                    'id': appt['appointment_id'],
                    'title': f"{appt['first_name']} {appt['last_name']} - Dr. {appt['doctor_name']}",
                    'start': appt_datetime.isoformat(),
                    'end': end_datetime.isoformat(),
                    'description': f"Patient: {appt['first_name']} {appt['last_name']}\nDoctor: Dr. {appt['doctor_name']}\nSpecialty: {appt['specialty']}",
                    'location': "MediCare Allergy & Wellness Center",
                    'status': appt.get('status', 'confirmed'),
                    'patient_email': appt.get('email', ''),
                    'patient_phone': appt.get('phone', ''),
                    'doctor_specialty': appt.get('specialty', '')
                }
                calendar_events.append(event)
            
            return {
                'success': True,
                'events': calendar_events,
                'total_appointments': len(calendar_events),
                'date_range': f"{start_date or 'All'} to {end_date or 'All'}"
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'Error retrieving calendar: {e}'
            }
    
    def export_full_calendar_excel(self, start_date: str = None, end_date: str = None) -> str:
        """Export full calendar to Excel with professional formatting"""
        try:
            calendar_data = self.get_all_appointments_calendar(start_date, end_date)
            
            if not calendar_data['success']:
                return None
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"medical_calendar_export_{timestamp}.xlsx"
            filepath = os.path.join(self.excel_export_path, filename)
            
            # Create workbook
            workbook = openpyxl.Workbook()
            ws = workbook.active
            ws.title = "Medical Calendar"
            
            # Styles
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = [
                "Appointment ID", "Date", "Time", "Patient Name", "Patient Email", 
                "Patient Phone", "Doctor", "Specialty", "Duration", "Status"
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            # Data rows
            for row, event in enumerate(calendar_data['events'], 2):
                start_time = datetime.fromisoformat(event['start'])
                end_time = datetime.fromisoformat(event['end'])
                duration = int((end_time - start_time).total_seconds() / 60)
                
                row_data = [
                    event['id'],
                    start_time.strftime('%Y-%m-%d'),
                    start_time.strftime('%H:%M'),
                    event['title'].split(' - ')[0],  # Patient name
                    event['patient_email'],
                    event['patient_phone'],
                    event['title'].split(' - ')[1] if ' - ' in event['title'] else '',  # Doctor
                    event['doctor_specialty'],
                    f"{duration} min",
                    event['status'].title()
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    
                    # Color code by status
                    if event['status'] == 'confirmed':
                        cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
                    elif event['status'] == 'pending':
                        cell.fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
                    elif event['status'] == 'cancelled':
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary
            summary_row = len(calendar_data['events']) + 3
            ws[f'A{summary_row}'] = f"Total Appointments: {calendar_data['total_appointments']}"
            ws[f'A{summary_row}'].font = Font(bold=True)
            
            workbook.save(filepath)
            return filename
            
        except Exception as e:
            print(f"Error exporting calendar to Excel: {e}")
            return None

# Integration functions for existing system
def integrate_with_medical_agent():
    """Integration helper for the medical agent system"""
    return CalendarIntegration()

if __name__ == "__main__":
    # Test the calendar integration
    calendar_system = CalendarIntegration()
    
    print("🗓️  Testing Calendar Integration System")
    print("=" * 50)
    
    # Test getting availability for a doctor
    doctor_id = "1"  # Assuming doctor ID 1 exists
    availability = calendar_system.get_doctor_availability(doctor_id)
    
    if 'error' not in availability:
        print(f"✅ Found {availability['total_slots']} available slots")
        print(f"Doctor: {availability['doctor']['doctor_name']}")
        print(f"Working Hours: {availability['working_hours']}")
    else:
        print(f"❌ Error: {availability['error']}")
    
    # Test calendar export
    export_file = calendar_system.export_full_calendar_excel()
    if export_file:
        print(f"✅ Calendar exported to: {export_file}")
    else:
        print("❌ Calendar export failed")
